import sys
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QLabel,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QFileDialog,
    QMessageBox,
    QComboBox,
    QLineEdit,
)
from PyQt5.QtGui import QIcon, QPixmap, QFont
from PyQt5 import QtCore
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5.QtGui import QPalette, QBrush, QPixmap
from PyQt5.QtCore import Qt
from openpyxl.styles import PatternFill
from openpyxl.worksheet.cell_range import CellRange
        

def update_rows_with_filter(source_file, target_file, target_sheet_name, filter_word):
    # Read the source Excel file
    df = pd.read_excel(source_file, sheet_name='Tüm Başvurular')

    # Filter rows based on the entered filter word in the 'Cinsiyet' column
    filtered_df = df[df['Cinsiyet'].str.contains(filter_word, case=False)]

    # Load the target workbook
    book = load_workbook(target_file)

    # Check if the target sheet exists in the workbook
    if target_sheet_name in book.sheetnames:
        # Get the existing target sheet
        sheet = book[target_sheet_name]
        # Clear the existing sheet data
        sheet.delete_rows(1, sheet.max_row)
        # Append the filtered data to the target sheet
        for row in dataframe_to_rows(filtered_df, index=False, header=True):
            sheet.append(row)
    else:
        # If the target sheet doesn't exist, create a new sheet and append the filtered data
        with pd.ExcelWriter(target_file, engine='openpyxl', mode='a') as writer:
            writer.book = book
            filtered_df.to_excel(writer, sheet_name=target_sheet_name, index=False)

    # Save the updated workbook
    book.save(target_file)

def update_rows_with_erkek(source_file, target_file, target_sheet_name):
    # Read the source Excel file
    df = pd.read_excel(source_file, sheet_name='Tüm Başvurular')

    # Filter rows with 'Erkek' in the 'Cinsiyet' column
    erkek_df = df[df['Cinsiyet'] == 'Erkek']

    # Load the target workbook
    book = load_workbook(target_file)

    # Check if the target sheet exists in the workbook
    if target_sheet_name in book.sheetnames:
        # Get the existing target sheet
        sheet = book[target_sheet_name]
        # Clear the existing sheet data
        sheet.delete_rows(1, sheet.max_row)
        # Append the filtered data to the target sheet
        for row in dataframe_to_rows(erkek_df, index=False, header=True):
            sheet.append(row)
    else:
        # If the target sheet doesn't exist, create a new sheet and append the filtered data
        with pd.ExcelWriter(target_file, engine='openpyxl', mode='a') as writer:
            writer.book = book
            erkek_df.to_excel(writer, sheet_name=target_sheet_name, index=False)

    # Save the updated workbook
    book.save(target_file)


class StyledComboBox(QComboBox):
    def __init__(self, parent=None):
        super(StyledComboBox, self).__init__(parent)
        # Set a custom stylesheet for the QComboBox
        self.setStyleSheet(
            """
            QComboBox {
                border: 2px solid #4CAF50; /* Green border */
                border-radius: 5px;
                padding: 5px;
                background-color: white;
                color: #333;
                font-size: 16px;
            }
            QComboBox:editable {
                background: white;
            }
            QComboBox:on { /* style when dropdown is open */
                border: 2px solid #45a049; /* Darker green border */
            }
            QComboBox QAbstractItemView {
                border: 2px solid #45a049; /* Darker green border */
                selection-background-color: #4CAF50; /* Green selection background */
                background-color: white;
                color: #333;
                font-size: 16px;
            }
            """
        )


class ExcelCopyApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Excel Copy System")
        self.setWindowIcon(
            QIcon("C:/Users/Muhammed Kartal/Desktop/Codes/WebApp/4_attendance_app/Logo/desktop-footer-logo@2x.png")
        )  # Replace with the correct path to your icon file
        self.setMinimumWidth(600)

        layout = QVBoxLayout()

        background_image_path = "C:/Users/Muhammed Kartal/Desktop/abstract-background-design-perfect-landing-page-background-other-you-want_288336-2.webp"
        palette = self.palette()
        palette.setBrush(QPalette.Background, QBrush(QPixmap(background_image_path).scaled(self.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation)))
        self.setPalette(palette)

        # Logo Label
        logo_label = QLabel(self)
        logo_pixmap = QPixmap("C:/Users/Muhammed Kartal/Desktop/Codes/WebApp/4_attendance_app/Logo/desktop-footer-logo@2x.png")
        logo_label.setPixmap(logo_pixmap.scaled(540, 550, aspectRatioMode=True))
        layout.addWidget(logo_label, alignment=QtCore.Qt.AlignHCenter)

         # Filter Word Section
        self.filter_label = QLabel("Filtreleme Kelimesi:", self)
        self.filter_word = QLineEdit(self)

        # Nested layout for Filter Word Section
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(self.filter_label)
        filter_layout.addWidget(self.filter_word)


        # Source File Section
        self.source_label = QLabel("Kaynak Excel Dosyası:", self)
        self.source_file_path = QLabel("", self)
        self.browse_source_button = QPushButton("Browse", self)
        self.browse_source_button.clicked.connect(self.select_source_file)

        # Nested layout for Source File Section
        source_layout = QHBoxLayout()
        source_layout.addWidget(self.source_label)
        source_layout.addWidget(self.source_file_path)
        source_layout.addWidget(self.browse_source_button, alignment=QtCore.Qt.AlignRight)

        # Target File Section
        self.target_label = QLabel("Hedef Excel Dosyası:", self)
        self.target_file_path = QLabel("", self)
        self.browse_target_button = QPushButton("Browse", self)
        self.browse_target_button.clicked.connect(self.select_target_file)

        # Nested layout for Target File Section
        target_layout = QHBoxLayout()
        target_layout.addWidget(self.target_label)
        target_layout.addWidget(self.target_file_path)
        target_layout.addWidget(self.browse_target_button, alignment=QtCore.Qt.AlignRight)

        # Target Sheet Section
        self.target_sheet_label = QLabel("Hedef Sheet Adı:", self)
        self.target_sheet_name = StyledComboBox(self)  # Use the custom StyledComboBox for target_sheet_name

        # Copy Button
        self.copy_button = QPushButton("Hedef Sayfayı Güncelle", self)
        self.copy_button.clicked.connect(self.update_target_sheet)

        font = QFont("Arial", 18)
        self.source_label.setFont(font)
        self.source_file_path.setFont(font)
        self.target_label.setFont(font)
        self.target_file_path.setFont(font)
        layout.addLayout(filter_layout)
        self.target_sheet_label.setFont(font)
        self.target_sheet_name.setFont(font)

        # Add all the widgets to the main layout
        layout.addLayout(source_layout)
        layout.addLayout(target_layout)
        layout.addWidget(self.target_sheet_label)
        layout.addWidget(self.target_sheet_name)
        layout.addWidget(self.copy_button, alignment=QtCore.Qt.AlignHCenter)

        # Reduce vertical spacing between the labels and file paths
        layout.setSpacing(10)

        # Apply the stylesheet to the main layout
        self.setStyleSheet(
            """
            QWidget {
                background-color: #F5F5F5; /* Light gray background */
                padding: 30px;
            }
            QLabel {
                font-size: 22px;
                color: #333;
                margin-bottom: 5px;
            }
            QPushButton {
                font-size: 16px;
                padding: 10px 30px;
                background-color: #4CAF50; /* Green button background */
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049; /* Darker green on hover */
            }
            """
        )

        self.setLayout(layout)

    
    def highlight_duplicates(self, target_file, target_sheet_name):
        # Read the target Excel file
        df = pd.read_excel(target_file, sheet_name=target_sheet_name)

        # Find duplicate columns
        duplicate_columns = df.columns[df.columns.duplicated(keep=False)]

        # Load the target workbook
        book = load_workbook(target_file)
        
        # Get the target sheet
        sheet = book[target_sheet_name]

        # Apply conditional formatting to highlight duplicate columns
        red_fill = PatternFill(start_color='FFFF0000',
                               end_color='FFFF0000',
                               fill_type='solid')
        for column in duplicate_columns:
            col_letter = get_column_letter(df.columns.get_loc(column) + 1)
            col_range = f"{col_letter}1:{col_letter}{sheet.max_row}"
            sheet.conditional_formatting.add(col_range, CellIsRule(operator='containsText',
                                                                   formula=[''],
                                                                   stopIfTrue=True,
                                                                   fill=red_fill))
            
        # Save the updated workbook
        book.save(target_file)

    def update_rows_with_erkek(source_file, target_file, target_sheet_name):
        # Read the source Excel file
        df = pd.read_excel(source_file, sheet_name='Sheet 1')

        # Filter rows with 'Erkek' in the 'Cinsiyet' column
        erkek_df = df[df['Cinsiyet'] == 'Erkek']

        # Load the target workbook
        book = load_workbook(target_file)

        # Check if the target sheet exists in the workbook
        if target_sheet_name in book.sheetnames:
            # Get the existing target sheet
            sheet = book[target_sheet_name]
            # Clear the existing sheet data
            sheet.delete_rows(1, sheet.max_row)
            # Append the filtered data to the target sheet
            for row in dataframe_to_rows(erkek_df, index=False, header=True):
                sheet.append(row)
        else:
            # If the target sheet doesn't exist, create a new sheet and append the filtered data
            with pd.ExcelWriter(target_file, engine='openpyxl', mode='a') as writer:
                writer.book = book
                erkek_df.to_excel(writer, sheet_name=target_sheet_name, index=False)

        # Save the updated workbook
        book.save(target_file)


    def select_source_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Select Source Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options
        )
        if file_name:
            self.source_file_path.setText(file_name)

    def select_target_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Select Target Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options
        )
        if file_name:
            self.target_file_path.setText(file_name)

            # Get the sheet names from the selected Excel file and add them to the target_sheet_name ComboBox
            sheets = get_sheet_names(file_name)
            self.target_sheet_name.clear()
            self.target_sheet_name.addItems(sheets)

    def update_target_sheet(self):
        source_file = self.source_file_path.text()
        target_file = self.target_file_path.text()
        target_sheet_name = self.target_sheet_name.currentText()    

        if not source_file or not target_file:
            QMessageBox.warning(
                self, "Warning", "Lütfen hem kaynak hem de hedef Excel dosyalarını seçin."
            )
            return

        filter_word = self.filter_word.text()  # Get the entered filter word

        try:
            update_rows_with_filter(source_file, target_file, target_sheet_name, filter_word)
            self.highlight_duplicates(target_file, target_sheet_name)  # Highlight duplicates after updating
            QMessageBox.information(
                self,
                "Başarı",
                f"'Cinsiyet' sütununda '{filter_word}' içeren satırlar hedef sayfada güncellendi ve yinelenen sütunlar vurgulandı.",
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


def get_sheet_names(file_path):
    sheets = pd.ExcelFile(file_path).sheet_names
    return sheets


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelCopyApp()
    window.show()
    sys.exit(app.exec_())
